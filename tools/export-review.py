#!/usr/bin/env python3
"""把 content-src 的词库与对话导出成一份可批注的审校 Excel。

用途：交给母语者/老师逐条过一遍，在「修改建议」列直接写改法。
改完后按 README 的流程回填 content-src/batches/*.json 再重新打包。

依赖：openpyxl（仅审校用，不是应用运行依赖）
用法：python3 tools/export-review.py [输出路径]
"""
import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "印尼语词库审校.xlsx"

FONT = "Calibri"
HEAD_FILL = PatternFill("solid", fgColor="2F6FED")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=11)
ID_FONT = Font(name=FONT, size=11, bold=True, color="1F4FB0")
NOTE_FILL = PatternFill("solid", fgColor="FFF7E0")
THIN = Side(style="thin", color="D6DEE9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_header(ws, headers):
    for i, (title, width) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def style_row(ws, row, ncols, note_cols, id_cols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = ID_FONT if col in id_cols else BODY_FONT
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if col in note_cols:
            c.fill = NOTE_FILL


def add_flag_validation(ws, col_letter, last_row):
    dv = DataValidation(
        type="list", formula1='"OK,要改,不确定"', allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def sheet_words(wb, packs):
    ws = wb.create_sheet("词条")
    write_header(
        ws,
        [
            ("包号", 7), ("主题", 11), ("副标题", 13), ("序", 5),
            ("印尼语", 20), ("词性", 9), ("中文释义", 22),
            ("例句", 34), ("例句翻译", 26),
            ("判定", 9), ("修改建议", 34),
        ],
    )
    row = 2
    for p in packs:
        for i, w in enumerate(p["words"], start=1):
            ws.cell(row=row, column=1, value=p["stage"])
            ws.cell(row=row, column=2, value=p["title"])
            ws.cell(row=row, column=3, value=p["subtitle"])
            ws.cell(row=row, column=4, value=i)
            ws.cell(row=row, column=5, value=w["word"])
            ws.cell(row=row, column=6, value=w["pos"])
            ws.cell(row=row, column=7, value=w["zh"])
            ws.cell(row=row, column=8, value=w["example"])
            ws.cell(row=row, column=9, value=w["exampleZh"])
            style_row(ws, row, 11, note_cols={10, 11}, id_cols={5})
            row += 1
    add_flag_validation(ws, "J", row - 1)
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:1"
    return row - 2


def sheet_dialogs(wb, dialogs):
    ws = wb.create_sheet("对话")
    write_header(
        ws,
        [
            ("场景", 12), ("场景（印尼语）", 18), ("行", 5), ("说话人", 8),
            ("印尼语", 40), ("中文", 30), ("判定", 9), ("修改建议", 34),
        ],
    )
    row = 2
    for d in dialogs:
        for i, l in enumerate(d["lines"], start=1):
            ws.cell(row=row, column=1, value=d["sceneZh"])
            ws.cell(row=row, column=2, value=d["scene"])
            ws.cell(row=row, column=3, value=i)
            ws.cell(row=row, column=4, value=l["speaker"])
            ws.cell(row=row, column=5, value=l["id_text"])
            ws.cell(row=row, column=6, value=l["zh"])
            style_row(ws, row, 8, note_cols={7, 8}, id_cols={5})
            row += 1
    add_flag_validation(ws, "G", row - 1)
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:1"
    return row - 2


def sheet_extras(wb, dialogs):
    ws = wb.create_sheet("关键句与生词")
    write_header(
        ws,
        [
            ("场景", 12), ("类型", 9), ("印尼语", 34), ("中文", 26),
            ("判定", 9), ("修改建议", 34),
        ],
    )
    row = 2
    for d in dialogs:
        for k in d["keyPhrases"]:
            ws.cell(row=row, column=1, value=d["sceneZh"])
            ws.cell(row=row, column=2, value="关键句")
            ws.cell(row=row, column=3, value=k["id_text"])
            ws.cell(row=row, column=4, value=k["zh"])
            style_row(ws, row, 6, note_cols={5, 6}, id_cols={3})
            row += 1
        for v in d["vocab"]:
            ws.cell(row=row, column=1, value=d["sceneZh"])
            ws.cell(row=row, column=2, value="生词")
            ws.cell(row=row, column=3, value=v["word"])
            ws.cell(row=row, column=4, value=v["zh"])
            style_row(ws, row, 6, note_cols={5, 6}, id_cols={3})
            row += 1
    add_flag_validation(ws, "E", row - 1)
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:1"
    return row - 2


def sheet_readme(wb, counts):
    ws = wb.create_sheet("说明", 0)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 82
    lines = [
        ("印尼语学习 · 内容审校表", ""),
        ("", ""),
        ("怎么用", "逐行读「印尼语」和「例句」。地道就在「判定」选 OK；"),
        ("", "有问题选「要改」，并在「修改建议」里写正确说法。拿不准选「不确定」。"),
        ("", ""),
        ("重点看", "1. 印尼语词是不是当地真在用的说法（不是书面直译）"),
        ("", "2. 例句是不是自然口语，语序、词缀有没有错"),
        ("", "3. 中文释义准不准，有没有漏掉常用义项"),
        ("", "4. 对话里的敬语（Bapak/Ibu/Mas/Mbak）用得对不对"),
        ("", ""),
        ("不用管", "配图、排版、颜色 —— 这些跟内容无关。"),
        ("", ""),
        ("规模", f"词条 {counts['words']} 条 / 对话 {counts['lines']} 行 / 关键句与生词 {counts['extras']} 条"),
        ("", ""),
        ("改完之后", "把这份表发回来，我按「修改建议」回填 content-src/batches/*.json，"),
        ("", "然后重新打包发布。原文件不用你手动改。"),
    ]
    for i, (a, b) in enumerate(lines, start=1):
        ca = ws.cell(row=i, column=1, value=a)
        cb = ws.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, bold=True, size=11)
        cb.font = BODY_FONT
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=16, color="2F6FED")


def main():
    skeleton = json.loads((ROOT / "content-src/skeleton.json").read_text(encoding="utf-8"))
    words = json.loads((ROOT / "content-src/words.json").read_text(encoding="utf-8"))
    packs = [{**p, "words": words.get(p["id"], [])} for p in skeleton if words.get(p["id"])]
    dialogs = json.loads((ROOT / "content-src/dialogs.json").read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    n_words = sheet_words(wb, packs)
    n_lines = sheet_dialogs(wb, dialogs)
    n_extras = sheet_extras(wb, dialogs)
    sheet_readme(wb, {"words": n_words, "lines": n_lines, "extras": n_extras})

    wb.save(OUT)
    print(f"已生成 {OUT}")
    print(f"  词条 {n_words} 行 / 对话 {n_lines} 行 / 关键句与生词 {n_extras} 行")


if __name__ == "__main__":
    main()
